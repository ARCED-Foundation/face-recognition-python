#!/usr/bin/env python3
"""
Face Recognition Script using DeepFace
Compares faces from a CSV file against an image database
"""

import argparse
import pandas as pd
import os
import sys
from datetime import datetime
from pathlib import Path
import numpy as np
from deepface import DeepFace
import warnings
warnings.filterwarnings('ignore')

def save_to_excel(df, output_path):
    """Save DataFrame to a well-formatted Excel file"""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Save to Excel
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Face Recognition Results', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Face Recognition Results']
            
            # Define styles
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            
            success_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            warning_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            error_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            
            high_score_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
            medium_score_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
            low_score_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
            
            border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            # Format header row
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border
            
            # Auto-adjust column widths and format cells
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                    
                    # Add borders to all cells
                    cell.border = border
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
                
                adjusted_width = min(max_length + 2, 50)  # Max width of 50
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Color code status column
            if 'status' in df.columns:
                status_col_idx = df.columns.get_loc('status') + 1
                for row in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row, column=status_col_idx)
                    if cell.value == 'success':
                        cell.fill = success_fill
                    elif cell.value == 'no_face_detected':
                        cell.fill = warning_fill
                    elif cell.value == 'image_not_found':
                        cell.fill = error_fill
            
            # Color code best_match_score column
            if 'best_match_score' in df.columns:
                score_col_idx = df.columns.get_loc('best_match_score') + 1
                for row in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row, column=score_col_idx)
                    try:
                        score = float(cell.value)
                        if score >= 80:
                            cell.fill = high_score_fill
                        elif score >= 60:
                            cell.fill = medium_score_fill
                        elif score > 0:
                            cell.fill = low_score_fill
                    except:
                        pass
            
            # Freeze the header row
            worksheet.freeze_panes = 'A2'
            
            # Add summary sheet
            summary_sheet = workbook.create_sheet('Summary')
            
            # Summary statistics
            total = len(df)
            successful = len(df[df['status'] == 'success']) if 'status' in df.columns else 0
            with_matches = len(df[df['matches_found'] > 0]) if 'matches_found' in df.columns else 0
            avg_score = df[df['status'] == 'success']['best_match_score'].mean() if 'best_match_score' in df.columns and successful > 0 else 0
            
            summary_data = [
                ['Face Recognition Summary', ''],
                ['', ''],
                ['Metric', 'Value'],
                ['Total Images Processed', total],
                ['Successfully Analyzed', successful],
                ['Images with Matches', with_matches],
                ['Average Best Match Score', f'{avg_score:.2f}%' if successful > 0 else 'N/A'],
                ['', ''],
                ['Status Breakdown', ''],
                ['Success', len(df[df['status'] == 'success']) if 'status' in df.columns else 0],
                ['No Face Detected', len(df[df['status'] == 'no_face_detected']) if 'status' in df.columns else 0],
                ['Image Not Found', len(df[df['status'] == 'image_not_found']) if 'status' in df.columns else 0],
            ]
            
            for row_idx, row_data in enumerate(summary_data, 1):
                for col_idx, value in enumerate(row_data, 1):
                    cell = summary_sheet.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    cell.alignment = Alignment(vertical='center')
                    
                    if row_idx == 1:
                        cell.font = Font(bold=True, size=14)
                    elif row_idx == 3 or row_idx == 9:
                        cell.fill = header_fill
                        cell.font = header_font
            
            summary_sheet.column_dimensions['A'].width = 30
            summary_sheet.column_dimensions['B'].width = 20
        
        print("✓ Excel file formatted successfully")
        
    except ImportError:
        print("Warning: openpyxl not installed. Saving as basic Excel without formatting.")
        df.to_excel(output_path, index=False)
    except Exception as e:
        print(f"Warning: Error formatting Excel: {e}")
        print("Saving as basic Excel file...")
        df.to_excel(output_path, index=False)

def find_faces_in_image(image_path, model_name='Facenet'):
    """Find and return face embeddings from an image"""
    try:
        # Extract face embeddings
        embedding = DeepFace.represent(
            img_path=image_path,
            model_name=model_name,
            enforce_detection=False,
            detector_backend='opencv'
        )
        
        if len(embedding) == 0:
            print(f"Warning: No face found in {image_path}")
            return None
        
        # Return the first face embedding found
        return np.array(embedding[0]['embedding'])
    except Exception as e:
        print(f"Error processing {image_path}: {e}")
        return None

def cosine_similarity(vec1, vec2):
    """Calculate cosine similarity between two vectors"""
    dot_product = np.dot(vec1, vec2)
    norm1 = np.linalg.norm(vec1)
    norm2 = np.linalg.norm(vec2)
    
    if norm1 == 0 or norm2 == 0:
        return 0
    
    similarity = dot_product / (norm1 * norm2)
    # Convert to 0-100 scale
    return max(0, min(100, (similarity + 1) * 50))

def compare_faces(source_embedding, db_embeddings, threshold=70):
    """
    Compare source face with database faces
    Returns list of matches with probabilities
    """
    matches = []
    
    for db_name, db_embedding in db_embeddings.items():
        if db_embedding is not None:
            # Calculate cosine similarity
            similarity = cosine_similarity(source_embedding, db_embedding)
            
            matches.append({
                'matched_image': db_name,
                'similarity_score': round(similarity, 2),
                'is_match': similarity >= threshold
            })
    
    # Sort by similarity score (highest first)
    matches.sort(key=lambda x: x['similarity_score'], reverse=True)
    return matches

def get_all_images(base_path, recursive=False, supported_formats=('.jpg', '.jpeg', '.png', '.bmp', '.gif')):
    """Get all image files from directory, optionally recursive"""
    images = []
    
    if recursive:
        for root, dirs, files in os.walk(base_path):
            for f in files:
                if f.lower().endswith(supported_formats):
                    # Store relative path from base_path
                    rel_path = os.path.relpath(os.path.join(root, f), base_path)
                    images.append(rel_path)
    else:
        for f in os.listdir(base_path):
            full_path = os.path.join(base_path, f)
            if os.path.isfile(full_path) and f.lower().endswith(supported_formats):
                images.append(f)
    
    return images

def main():
    parser = argparse.ArgumentParser(
        description='Face Recognition: Compare faces from CSV against image database',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Example usage:
  python3 face-recognize.py -source data.csv -imageDB ./images -output results.csv
  python3 face-recognize.py -source data.csv -imageDB ./images -output ./output/${today}-results.csv -threshold 75
  python3 face-recognize.py -source data.csv -imageDB ./images -output results.csv -recursive
        """
    )
    
    parser.add_argument('-source', required=True, 
                       help='Path to source CSV file (must contain "image" column)')
    parser.add_argument('-imageDB', required=True,
                       help='Path to image database folder')
    parser.add_argument('-output', required=True,
                       help='Path to output CSV file (use ${today} for current date)')
    parser.add_argument('-threshold', type=float, default=70,
                       help='Face match threshold (0-100, default: 70, higher = stricter)')
    parser.add_argument('-top', type=int, default=5,
                       help='Number of top matches to include (default: 5)')
    parser.add_argument('-recursive', action='store_true',
                       help='Search for images recursively in all subfolders')
    parser.add_argument('-model', default='Facenet',
                       choices=['VGG-Face', 'Facenet', 'Facenet512', 'OpenFace', 'DeepFace', 'DeepID', 'ArcFace', 'Dlib', 'SFace'],
                       help='Face recognition model to use (default: Facenet)')
    parser.add_argument('-excel', action='store_true',
                       help='Output as formatted Excel file (.xlsx) instead of CSV')
    
    args = parser.parse_args()
    
    # Replace ${today} with current date
    today = datetime.now().strftime('%Y-%m-%d')
    output_path = args.output.replace('${today}', today)
    
    # Auto-adjust extension for Excel output
    if args.excel and not output_path.lower().endswith('.xlsx'):
        output_path = os.path.splitext(output_path)[0] + '.xlsx'
    
    # Validate paths
    if not os.path.exists(args.source):
        print(f"Error: Source CSV file not found: {args.source}")
        sys.exit(1)
    
    if not os.path.exists(args.imageDB):
        print(f"Error: Image database folder not found: {args.imageDB}")
        sys.exit(1)
    
    # Create output directory if it doesn't exist
    output_dir = os.path.dirname(output_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"Created output directory: {output_dir}")
    
    print("=" * 60)
    print("Face Recognition Script (DeepFace)")
    print("=" * 60)
    print(f"Source CSV: {args.source}")
    print(f"Image Database: {args.imageDB}")
    print(f"Output CSV: {output_path}")
    print(f"Match Threshold: {args.threshold}")
    print(f"Top Matches: {args.top}")
    print(f"Recursive Search: {args.recursive}")
    print(f"Model: {args.model}")
    print(f"Output Format: {'Excel (.xlsx)' if args.excel else 'CSV (.csv)'}")
    print("=" * 60)
    
    # Load CSV
    try:
        df = pd.read_csv(args.source)
        if 'image' not in df.columns:
            print("Error: CSV file must contain an 'image' column")
            sys.exit(1)
        print(f"✓ Loaded CSV with {len(df)} rows")
    except Exception as e:
        print(f"Error reading CSV: {e}")
        sys.exit(1)
    
    # Load all images from database
    print("\nLoading image database...")
    supported_formats = ('.jpg', '.jpeg', '.png', '.bmp', '.gif')
    db_images = get_all_images(args.imageDB, args.recursive, supported_formats)
    
    print(f"Found {len(db_images)} images in database")
    if args.recursive:
        print("(Including all subfolders)")
    
    # Pre-load all database image embeddings
    print(f"\nEncoding database images using {args.model}...")
    db_embeddings = {}
    for i, img_name in enumerate(db_images, 1):
        img_path = os.path.join(args.imageDB, img_name)
        print(f"  [{i}/{len(db_images)}] Encoding {img_name}...", end='\r')
        db_embeddings[img_name] = find_faces_in_image(img_path, args.model)
    
    print(f"\n✓ Encoded {sum(1 for v in db_embeddings.values() if v is not None)} faces")
    
    # Process each row in CSV
    print("\nProcessing source images...")
    results = []
    
    for idx, row in df.iterrows():
        image_name = row['image']
        print(f"\n[{idx + 1}/{len(df)}] Processing: {image_name}")
        
        # Handle both absolute and relative paths
        if os.path.isabs(image_name):
            image_path = image_name
        else:
            image_path = os.path.join(args.imageDB, image_name)
        
        # Check if image exists
        if not os.path.exists(image_path):
            print(f"  Warning: Image not found: {image_name}")
            row_result = row.to_dict()
            row_result['status'] = 'image_not_found'
            row_result['best_match'] = 'N/A'
            row_result['best_match_score'] = 0.0
            results.append(row_result)
            continue
        
        # Load and encode source image
        source_embedding = find_faces_in_image(image_path, args.model)
        
        if source_embedding is None:
            print(f"  Warning: No face detected in {image_name}")
            row_result = row.to_dict()
            row_result['status'] = 'no_face_detected'
            row_result['best_match'] = 'N/A'
            row_result['best_match_score'] = 0.0
            results.append(row_result)
            continue
        
        # Compare with database (excluding self)
        db_embeddings_filtered = {}
        for k, v in db_embeddings.items():
            # Normalize paths for comparison
            k_normalized = os.path.normpath(k)
            image_name_normalized = os.path.normpath(image_name)
            
            if k_normalized != image_name_normalized and v is not None:
                db_embeddings_filtered[k] = v
        
        matches = compare_faces(source_embedding, db_embeddings_filtered, args.threshold)
        
        # Prepare result row
        row_result = row.to_dict()
        row_result['status'] = 'success'
        row_result['total_comparisons'] = len(db_embeddings_filtered)
        row_result['matches_found'] = sum(1 for m in matches if m['is_match'])
        
        if matches:
            # Best match
            row_result['best_match'] = matches[0]['matched_image']
            row_result['best_match_score'] = matches[0]['similarity_score']
            
            # Top N matches
            for i, match in enumerate(matches[:args.top], 1):
                row_result[f'match_{i}_image'] = match['matched_image']
                row_result[f'match_{i}_score'] = match['similarity_score']
            
            print(f"  ✓ Best match: {matches[0]['matched_image']} "
                  f"(score: {matches[0]['similarity_score']}%)")
        else:
            row_result['best_match'] = 'N/A'
            row_result['best_match_score'] = 0.0
            print(f"  No matches found")
        
        results.append(row_result)
    
    # Create output DataFrame
    output_df = pd.DataFrame(results)
    
    # Save to file
    try:
        if args.excel:
            # Save as formatted Excel file
            save_to_excel(output_df, output_path)
        else:
            # Save as CSV
            output_df.to_csv(output_path, index=False)
        
        print(f"\n{'=' * 60}")
        print(f"✓ Results saved to: {output_path}")
        print(f"{'=' * 60}")
        
        # Summary statistics
        total = len(output_df)
        successful = len(output_df[output_df['status'] == 'success'])
        with_matches = len(output_df[output_df['matches_found'] > 0])
        
        print(f"\nSummary:")
        print(f"  Total images processed: {total}")
        print(f"  Successfully analyzed: {successful}")
        print(f"  Images with matches: {with_matches}")
        
        if successful > 0:
            avg_score = output_df[output_df['status'] == 'success']['best_match_score'].mean()
            print(f"  Average best match score: {avg_score:.2f}%")
        
    except Exception as e:
        print(f"Error saving results: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()